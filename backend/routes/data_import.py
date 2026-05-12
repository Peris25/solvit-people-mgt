"""Data Import — Excel templates download, validation, bulk import, history log."""
from fastapi import APIRouter, HTTPException, Request, UploadFile, File, Form
from fastapi.responses import StreamingResponse, FileResponse
from datetime import datetime, timezone
from pathlib import Path
from database import get_db
from utils.auth import get_current_user
import uuid
import io
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

router = APIRouter(prefix="/data-import", tags=["data_import"])

UPLOAD_ROOT = Path(os.environ.get("IMPORT_UPLOAD_ROOT", "/app/backend/uploads/imports"))
UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)

TEMPLATES = {
    "fte_employee": {
        "title": "FTE Employee Import",
        "columns": [
            ("Employee ID",      "Unique ID. Leave blank to auto-generate."),
            ("Full Name",        "First + last name as it appears on official records."),
            ("Job Title",        "e.g. Inspections Analyst, Finance Manager."),
            ("Department",       "Must match an active department from Masters Settings."),
            ("Job Level (L1–L5)","One of: L1, L2, L3, L4, L5."),
            ("Employment Start Date", "DD/MM/YYYY format."),
            ("Employment Type",  "Full-Time / Part-Time / Contract."),
            ("Probation End Date","DD/MM/YYYY. Leave blank if not on probation."),
            ("Line Manager Full Name", "Existing employee. Used to look up line_manager_id."),
            ("Work Email",       "Unique work email. Used for login."),
            ("Phone Number",     "Mobile in +254 format."),
            ("National ID Number","Kenyan National ID."),
            ("Date of Birth",    "DD/MM/YYYY."),
            ("Gender",           "Male / Female / Prefer not to say."),
            ("NSSF Number",      ""),
            ("NHIF Number",      ""),
            ("KRA PIN",          ""),
            ("Bank Name",        ""),
            ("Bank Branch",      ""),
            ("Bank Account Number", ""),
            ("Salary (KES)",     "Monthly gross in KES. Numbers only."),
            ("Pay Frequency",    "Monthly (default)."),
            ("Employment Status","Active / On Leave / Terminated."),
        ],
        "sample": [
            "EMP-NEW-001", "Jane Doe", "Inspections Analyst", "Operations", "L2",
            "01/02/2026", "Full-Time", "01/05/2026", "Sarah Njoroge",
            "jane.doe@solvit.co.ke", "+254700111222", "29384756", "15/06/1994", "Female",
            "NSSF-99887", "NHIF-44556", "A012345678X", "Equity Bank", "Westlands",
            "1234567890", "85000", "Monthly", "Active"
        ],
        "valid_values": {
            "Job Level (L1–L5)": ["L1","L2","L3","L4","L5"],
            "Employment Type": ["Full-Time","Part-Time","Contract"],
            "Employment Status": ["Active","On Leave","Terminated"],
            "Gender": ["Male","Female","Prefer not to say"],
            "Pay Frequency": ["Monthly"],
        }
    },
    "solver": {
        "title": "Solver Import",
        "columns": [
            ("Solver ID",        "Unique. Leave blank to auto-generate."),
            ("Full Name",        ""),
            ("Phone Number",     "+254 format. Used for SMS & MPesa."),
            ("Email",            "Optional."),
            ("National ID",      ""),
            ("Date of Onboarding","DD/MM/YYYY."),
            ("App Status",       "Active / Inactive / Suspended."),
            ("County",           "e.g. Nairobi."),
            ("Sub-County",       "e.g. Westlands."),
            ("Vehicle Types Qualified For","Comma-separated. e.g. Saloon, SUV."),
            ("Completed Jobs Total","Integer."),
            ("Accuracy Rate (%)","0–100."),
            ("Average Client Rating","0–5.0."),
            ("Current Performance Tier","Tier 1 / Tier 2 / Tier 3."),
        ],
        "sample": [
            "SOL-NEW-001","Peter Otieno","+254711333444","peter@example.com",
            "30192837","10/01/2026","Active","Nairobi","Westlands",
            "Saloon, SUV","42","92","4.7","Tier 1",
        ],
        "valid_values": {
            "App Status": ["Active","Inactive","Suspended"],
            "Current Performance Tier": ["Tier 1","Tier 2","Tier 3"],
        }
    },
    "historical_performance": {
        "title": "Historical Performance Data Import",
        "columns": [
            ("Employee ID",      "Must match existing employee."),
            ("Review Period",    "e.g. Q1 2025 / H1 2025 / 2024."),
            ("Review Date",      "DD/MM/YYYY."),
            ("Section A Score (1–3)", "1.00–3.00."),
            ("Section B Score (1–3)", "1.00–3.00."),
            ("Overall Score (1–3)",   "1.00–3.00. Optional — auto-calc when blank."),
            ("Performance Level","Exceeds / Meets / Below."),
            ("Reviewer Name",    "Free text."),
            ("Notes",            "Optional commentary."),
        ],
        "sample": [
            "EMP-001","Q4 2025","31/12/2025","1.2","1.4","1.3","Exceeds","Sarah Njoroge","Outstanding quarter."
        ],
        "valid_values": {
            "Performance Level": ["Exceeds","Meets","Below"],
        }
    }
}


def _build_template(kind: str) -> io.BytesIO:
    tpl = TEMPLATES[kind]
    wb = Workbook()
    # Sheet 1: Data
    ws = wb.active
    ws.title = "Data"
    headers = [c[0] for c in tpl["columns"]]
    tooltips = [c[1] for c in tpl["columns"]]
    header_fill = PatternFill(start_color="FF353F", end_color="FF353F", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if tooltips[i - 1]:
            cell.comment = Comment(tooltips[i - 1], "Solvit")
        ws.column_dimensions[cell.column_letter].width = max(18, min(40, len(h) + 5))
    # Sample row (greyed)
    sample_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    sample_font = Font(color="9CA3AF", italic=True)
    for i, v in enumerate(tpl["sample"], start=1):
        cell = ws.cell(row=2, column=i, value=v)
        cell.fill = sample_fill
        cell.font = sample_font
    ws.row_dimensions[1].height = 32

    # Sheet 2: Read Me
    rm = wb.create_sheet("Read Me")
    rm.column_dimensions["A"].width = 32
    rm.column_dimensions["B"].width = 80
    rm["A1"] = f"{tpl['title']} — Field Reference"
    rm["A1"].font = Font(size=14, bold=True, color="191919")
    rm["A2"] = "Row 2 of the Data tab shows a sample row (greyed out). Delete row 2 before importing real data."
    rm["A2"].font = Font(italic=True, color="525252")
    rm.append([])
    rm.append(["Field", "Description / Valid Values"])
    rm.cell(row=rm.max_row, column=1).font = Font(bold=True)
    rm.cell(row=rm.max_row, column=2).font = Font(bold=True)
    for col_name, tip in tpl["columns"]:
        valid = tpl["valid_values"].get(col_name)
        desc = tip
        if valid:
            desc = (desc + "  · Valid values: " + " / ".join(valid)) if desc else "Valid values: " + " / ".join(valid)
        rm.append([col_name, desc or "—"])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


@router.get("/template/{kind}")
async def download_template(kind: str, request: Request):
    user = await get_current_user(request)
    if user.get("role") not in ("hr_admin", "hr_manager", "it_admin"):
        raise HTTPException(status_code=403, detail="HR/IT Admin only")
    if kind not in TEMPLATES:
        raise HTTPException(status_code=404, detail="Unknown template kind")
    bio = _build_template(kind)
    fname = f"solvit_{kind}_template.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


def _norm_date(v):
    if not v:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None  # signal invalid


async def _validate_rows(kind: str, rows: list, db) -> list:
    """Return list of dicts: { row_index, raw, valid: bool, errors: [str], normalized: {...} }."""
    tpl = TEMPLATES[kind]
    headers = [c[0] for c in tpl["columns"]]
    valid_lookup = tpl["valid_values"]
    out = []

    # Existing IDs to detect duplicates
    if kind == "fte_employee":
        existing_ids = {e.get("employee_code") for e in await db.employees.find({}, {"employee_code": 1}).to_list(1000)}
        existing_emails = {e.get("work_email") for e in await db.employees.find({}, {"work_email": 1}).to_list(1000)}
    elif kind == "solver":
        existing_ids = {s.get("solver_code") for s in await db.solvers.find({}, {"solver_code": 1}).to_list(1000)}
        existing_emails = set()
    else:
        existing_ids = set()
        existing_emails = set()

    required = {
        "fte_employee": ["Full Name", "Job Title", "Department", "Job Level (L1–L5)", "Employment Start Date", "Work Email"],
        "solver": ["Full Name", "Phone Number", "Date of Onboarding"],
        "historical_performance": ["Employee ID", "Review Period", "Review Date", "Section A Score (1–3)", "Section B Score (1–3)"],
    }[kind]
    date_fields = {
        "fte_employee": ["Employment Start Date", "Probation End Date", "Date of Birth"],
        "solver": ["Date of Onboarding"],
        "historical_performance": ["Review Date"],
    }[kind]

    for idx, row in enumerate(rows, start=1):
        record = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        errors = []
        normalized = dict(record)

        for f in required:
            if not record.get(f) or str(record.get(f)).strip() == "":
                errors.append(f"`{f}` is required")

        for f in date_fields:
            v = record.get(f)
            if v:
                nd = _norm_date(v)
                if nd is None:
                    errors.append(f"`{f}` is not a valid date (use DD/MM/YYYY)")
                else:
                    normalized[f] = nd

        for f, allowed in valid_lookup.items():
            v = record.get(f)
            if v and str(v).strip() not in allowed:
                errors.append(f"`{f}` value '{v}' is not one of: {', '.join(allowed)}")

        # Duplicate
        id_field = {"fte_employee": "Employee ID", "solver": "Solver ID", "historical_performance": None}.get(kind)
        if id_field:
            rid = record.get(id_field)
            if rid and str(rid).strip() in existing_ids:
                errors.append(f"`{id_field}` '{rid}' already exists in the system")
        if kind == "fte_employee":
            email = record.get("Work Email")
            if email and email in existing_emails:
                errors.append(f"`Work Email` '{email}' already exists in the system")

        out.append({
            "row_index": idx,
            "raw": record,
            "normalized": normalized,
            "valid": not errors,
            "errors": errors,
        })
    return out


def _read_xlsx(content: bytes) -> list:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb["Data"] if "Data" in wb.sheetnames else wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) <= 1:
        return []
    # Skip header row; skip sample row when it looks like the canned sample (italic/grey isn't readable here, so we
    # treat as data — caller can keep it if they intended it).
    return [r for r in all_rows[1:] if any(c is not None and str(c).strip() != "" for c in r)]


@router.post("/validate")
async def validate_import(request: Request, kind: str = Form(...), file: UploadFile = File(...)):
    user = await get_current_user(request)
    if user.get("role") not in ("hr_admin", "hr_manager"):
        raise HTTPException(status_code=403, detail="HR Admin only")
    if kind not in TEMPLATES:
        raise HTTPException(status_code=400, detail="Unknown import kind")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")
    try:
        rows = _read_xlsx(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {e}")
    db = get_db()
    results = await _validate_rows(kind, rows, db)
    # Cache for the subsequent import call
    cache_id = str(uuid.uuid4())
    cache_path = UPLOAD_ROOT / f"{cache_id}.xlsx"
    with open(cache_path, "wb") as f:
        f.write(content)
    return {
        "cache_id": cache_id,
        "kind": kind,
        "filename": file.filename,
        "total_rows": len(results),
        "valid_count": sum(1 for r in results if r["valid"]),
        "error_count": sum(1 for r in results if not r["valid"]),
        "rows": results,
    }


@router.post("/import")
async def execute_import(request: Request):
    user = await get_current_user(request)
    if user.get("role") not in ("hr_admin", "hr_manager"):
        raise HTTPException(status_code=403, detail="HR Admin only")
    body = await request.json()
    kind = body.get("kind")
    cache_id = body.get("cache_id")
    only_valid = bool(body.get("only_valid", True))
    if kind not in TEMPLATES or not cache_id:
        raise HTTPException(status_code=400, detail="Missing kind or cache_id")
    cache_path = UPLOAD_ROOT / f"{cache_id}.xlsx"
    if not cache_path.exists():
        raise HTTPException(status_code=404, detail="Validation cache expired — re-upload the file")
    db = get_db()
    rows = _read_xlsx(cache_path.read_bytes())
    results = await _validate_rows(kind, rows, db)

    imported, skipped = 0, 0
    now = datetime.now(timezone.utc).isoformat()
    for r in results:
        if not r["valid"]:
            skipped += 1
            continue
        rec = r["normalized"]
        try:
            if kind == "fte_employee":
                doc = {
                    "id": str(uuid.uuid4()),
                    "tenant_id": "solvit",
                    "employee_code": rec.get("Employee ID") or f"EMP-{str(uuid.uuid4())[:6].upper()}",
                    "full_name": rec.get("Full Name"),
                    "role_title": rec.get("Job Title"),
                    "department": rec.get("Department"),
                    "job_level": rec.get("Job Level (L1–L5)"),
                    "start_date": rec.get("Employment Start Date"),
                    "employment_type": rec.get("Employment Type") or "Full-Time",
                    "probation_end_date": rec.get("Probation End Date"),
                    "work_email": rec.get("Work Email"),
                    "phone_number": rec.get("Phone Number"),
                    "national_id": rec.get("National ID Number"),
                    "date_of_birth": rec.get("Date of Birth"),
                    "gender": rec.get("Gender"),
                    "nssf_number": rec.get("NSSF Number"),
                    "nhif_number": rec.get("NHIF Number"),
                    "kra_pin": rec.get("KRA PIN"),
                    "bank_name": rec.get("Bank Name"),
                    "bank_branch": rec.get("Bank Branch"),
                    "bank_account": rec.get("Bank Account Number"),
                    "salary_kes": float(rec.get("Salary (KES)") or 0) if rec.get("Salary (KES)") else None,
                    "pay_frequency": rec.get("Pay Frequency") or "Monthly",
                    "lifecycle_state": (rec.get("Employment Status") or "Active"),
                    "line_manager_name": rec.get("Line Manager Full Name"),
                    "imported_at": now,
                    "imported_by": user["id"],
                }
                await db.employees.insert_one(doc)
                imported += 1
            elif kind == "solver":
                doc = {
                    "id": str(uuid.uuid4()),
                    "tenant_id": "solvit",
                    "solver_code": rec.get("Solver ID") or f"SOL-{str(uuid.uuid4())[:6].upper()}",
                    "full_name": rec.get("Full Name"),
                    "phone_number": rec.get("Phone Number"),
                    "email": rec.get("Email"),
                    "national_id": rec.get("National ID"),
                    "onboarded_at": rec.get("Date of Onboarding"),
                    "lifecycle_state": (rec.get("App Status") or "Active"),
                    "county": rec.get("County"),
                    "sub_county": rec.get("Sub-County"),
                    "vehicle_categories": [v.strip() for v in str(rec.get("Vehicle Types Qualified For") or "").split(",") if v.strip()],
                    "jobs_completed": int(rec.get("Completed Jobs Total") or 0),
                    "accuracy_score": float(rec.get("Accuracy Rate (%)") or 0),
                    "client_rating_average": float(rec.get("Average Client Rating") or 0),
                    "performance_tier": rec.get("Current Performance Tier"),
                    "imported_at": now,
                    "imported_by": user["id"],
                }
                await db.solvers.insert_one(doc)
                imported += 1
            else:  # historical_performance
                doc = {
                    "id": str(uuid.uuid4()),
                    "tenant_id": "solvit",
                    "employee_id": rec.get("Employee ID"),
                    "cycle_label": rec.get("Review Period"),
                    "cycle_year": int(str(rec.get("Review Period") or "")[-4:]) if str(rec.get("Review Period") or "")[-4:].isdigit() else None,
                    "review_date": rec.get("Review Date"),
                    "section_a_score": float(rec.get("Section A Score (1–3)") or 0),
                    "section_b_score": float(rec.get("Section B Score (1–3)") or 0),
                    "overall_score": float(rec.get("Overall Score (1–3)")) if rec.get("Overall Score (1–3)") else None,
                    "rating": rec.get("Performance Level"),
                    "reviewer_name": rec.get("Reviewer Name"),
                    "notes": rec.get("Notes"),
                    "status": "Imported",
                    "imported_at": now,
                    "imported_by": user["id"],
                }
                await db.performance_reviews.insert_one(doc)
                imported += 1
        except Exception as ex:
            skipped += 1
            print(f"[data_import] row {r['row_index']} failed: {ex}")

    # Log import history
    log_id = str(uuid.uuid4())
    log = {
        "id": log_id,
        "tenant_id": "solvit",
        "kind": kind,
        "filename": body.get("filename") or f"{cache_id}.xlsx",
        "rows_total": len(results),
        "rows_imported": imported,
        "rows_skipped": skipped,
        "only_valid": only_valid,
        "stored_path": str(cache_path),
        "imported_at": now,
        "imported_by": user["id"],
        "imported_by_name": user.get("full_name") or user.get("email"),
    }
    await db.data_import_log.insert_one(log)
    return {"import_log_id": log_id, "imported": imported, "skipped": skipped, "timestamp": now}


@router.get("/history")
async def import_history(request: Request, limit: int = 50):
    user = await get_current_user(request)
    if user.get("role") not in ("hr_admin", "hr_manager", "it_admin"):
        raise HTTPException(status_code=403, detail="HR/IT Admin only")
    db = get_db()
    rows = await db.data_import_log.find({"tenant_id": "solvit"}).sort("imported_at", -1).to_list(min(limit, 200))
    out = []
    for r in rows:
        r.pop("_id", None)
        out.append(r)
    return out


@router.get("/history/{log_id}/download")
async def download_original_import(log_id: str, request: Request):
    user = await get_current_user(request)
    if user.get("role") not in ("hr_admin", "hr_manager", "it_admin"):
        raise HTTPException(status_code=403, detail="HR/IT Admin only")
    db = get_db()
    log = await db.data_import_log.find_one({"id": log_id})
    if not log:
        raise HTTPException(status_code=404, detail="Log entry not found")
    p = Path(log.get("stored_path") or "")
    if not p.exists():
        raise HTTPException(status_code=404, detail="Original file no longer available")
    return FileResponse(p, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        filename=log.get("filename") or p.name)


@router.post("/error-report")
async def error_report_xlsx(request: Request):
    """Generate an Excel error report from a previously-cached validation result."""
    user = await get_current_user(request)
    if user.get("role") not in ("hr_admin", "hr_manager"):
        raise HTTPException(status_code=403, detail="HR Admin only")
    body = await request.json()
    rows = body.get("rows") or []
    error_rows = [r for r in rows if not r.get("valid")]
    wb = Workbook()
    ws = wb.active
    ws.title = "Errors"
    ws.append(["Row", "Errors", "Field Values"])
    for row in error_rows:
        ws.append([row["row_index"], " · ".join(row.get("errors") or []),
                   " | ".join(f"{k}={v}" for k, v in (row.get("raw") or {}).items() if v)])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=solvit_import_errors.xlsx"})
